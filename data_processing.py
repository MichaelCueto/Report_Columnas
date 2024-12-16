import pandas as pd
import numpy as np
import math
import warnings
import os
import re
from openpyxl import load_workbook
from datetime import datetime

# Ignorar todas las advertencias
warnings.filterwarnings("ignore")

class ExcelReader:
    def __init__(self, folder_path):
        self.folder_path = folder_path

    def read_excel_files(self):
        file_names = [f for f in os.listdir(self.folder_path) if f.endswith('.xlsx')]
        patterns = [
            (r'18513 - Enriquecido TCS - (\d{3}) 1M-', r'TCS \1'),
            (r'18513 - Enriquecido TCC - (\d{3}) 8M-', r'TCC \1'),
            (r'18513 - Mixto TCS - (\d{3}) 1M-', r'TCS \1'),
            (r'18513 - Transicional TCS - (\d{3}) 1M-', r'TCS \1'),
            (r'18513 Composito Enriquecido TBCE-(\d{2}) 1M -', r'TBCE\1 OP Enriquecido'),
            (r'18513 Composito Mixto TBCM-(\d{2}) 1M -', r'TBCM\1 OP Mixto'),
            (r'18513 Composito GEOTECNIA', r'Geotecnia - ROM3'),
            (r'18513 Composito ROM (\d{1}) -', r'ROM\1'),
            (r'TBC(\d{3})', r'TBC \1'),
            (r'TBC(\d{3}) - D', r'TBC \1 - D'),
            (r'TBC(\d{3}) - 3D', r'TBC \1 - 3D')
        ]
        dfs = []
        for file_name in file_names:
            # Quitar la extensión .xlsx
            file_name_no_ext = os.path.splitext(file_name)[0]
            original_name = file_name_no_ext
            for pattern, replacement in patterns:
                file_name_no_ext = re.sub(pattern, replacement, file_name_no_ext)
            # Eliminar ceros a la izquierda en los números después de 'TCS ' o 'TBCM'
            file_name_no_ext = file_name_no_ext.replace('_', ' ')
            file_name_no_ext = re.sub(r'TCS 0*', 'TCS ', file_name_no_ext)
            file_name_no_ext = re.sub(r'TCC 0*', 'TCC ', file_name_no_ext)
            file_name_no_ext = re.sub(r'TBC 0*', 'TBC ', file_name_no_ext)
            file_name_no_ext = re.sub(r'TBCM0*', 'TBCM', file_name_no_ext)
            df = self.read_single_excel(os.path.join(self.folder_path,file_name))
            df['Nombre del Archivo'] = file_name_no_ext
            dfs.append(df)
        return dfs

    def read_single_excel(self, file_path):
        try:
            # Determinar la hoja según el nombre del archivo
            if 'enriquecido' in file_path.lower():
                sheet_name = 'ENRIQUECIDO'
            elif 'transicional' in file_path.lower():
                sheet_name = 'TRANSICIONAL'
            elif 'mixto' in file_path.lower():
                sheet_name = 'MIXTO'
            else:
                sheet_name = 'COLUMNA'
            df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=28, usecols='C:Z')
            last_row_col_5 = df[df.iloc[:, 6].notna()].index[-1]
            last_row_col_8 = df[df.iloc[:, 7].notna()].index[-1]
            last_row = max(last_row_col_5, last_row_col_8)
            df =  df.iloc[:last_row + 1]
            df1 = df.drop(df.columns[2], axis=1)
            df1 = df1.fillna(0)
            inicio_df = df1.index[0]
            df0 = df.head(inicio_df)
            df0 = df0.drop(df0.columns[2], axis=1)
            df_concatenado = pd.concat([df0, df1], axis=0)
            df_concatenado.columns = ['Tiempo', 'Fecha', 'Día', 'Peso(g)_PLS', 'ÁcidoLibre(g/Kg)_PLS', 'pH_PLS', 'ORP(mV)_PLS', 'Temp_PLS', 'Temp_Amb_PLS', 'Temp_Column_PLS', 'Cu(g/Kg)_PLS',
                                      'Fe(g/Kg)_PLS', 'Fe+2(g/Kg)_PLS', 'Extraccion_Cu%', 'H2SO4_Total(Kg/T)', 'Peso(g)_Feed', 'ÁcidoLibre(g/Kg)_Feed', 'pH_Feed', 'ORP(mV)_Feed', 'Acido_ingresa(g)',
                                      'Cu(g/Kg)_Feed', 'Fe(g/Kg)_Feed', 'Fe+2(g/Kg)_Feed']
            df_concatenado['Nombre del Archivo'] = os.path.basename(file_path)
            return df_concatenado
        except Exception as e:
            print(f"Error al procesar el archivo {file_path}: {e}")
            return None
        
    def calculate_parameters(self):
        results = []
        file_names = [f for f in os.listdir(self.folder_path) if f.endswith('.xlsx')]
        for file_name in file_names:
            file_path = os.path.join(self.folder_path, file_name)
            try:
                if 'enriquecido' in file_name.lower():
                    sheet_name = 'ENRIQUECIDO'
                elif 'transicional' in file_name.lower():
                    sheet_name = 'TRANSICIONAL'
                elif 'mixto' in file_name.lower():
                    sheet_name = 'MIXTO'
                else:
                    sheet_name = 'COLUMNA'
                dry_column_charge = float(load_workbook(file_path)[sheet_name]['AI51'].value)
                column_size = load_workbook(file_path)[sheet_name]['D13'].value
                peso_inicial = float(load_workbook(file_path)['Mineral']['L16'].value)
                values = load_workbook(file_path)['Mineral']['J24'].value.split(',')
                cu_inicial = float(values[0].split('=')[-1].strip().split()[0])
                #Acido = float(values[1].split('=')[-1].strip().split()[0])
                fe_total_inicial = float(values[2].split('=')[-1].strip().split()[0])
                fe_2_inicial = float(values[3].split('=')[-1].strip().split()[0])
                h2so4_inicial = float(load_workbook(file_path)['Mineral']['L29'].value)
                altura_inicial_mineral = load_workbook(file_path)[sheet_name]['AI53'].value
                cu_total = load_workbook(file_path)[sheet_name]['AG35'].value
                fe_total = load_workbook(file_path)[sheet_name]['AH35'].value
                cu_acido = load_workbook(file_path)[sheet_name]['AJ35'].value
                cu_cn = load_workbook(file_path)[sheet_name]['AK35'].value
                cu_residual = load_workbook(file_path)[sheet_name]['AL35'].value
                material_moisture = load_workbook(file_path)[sheet_name]['AI39'].value
                results.append((dry_column_charge,column_size,peso_inicial,cu_inicial,h2so4_inicial,fe_total_inicial,fe_2_inicial,altura_inicial_mineral, cu_total, fe_total,cu_acido,cu_cn,cu_residual, material_moisture))
            except Exception as e:
                print(f"Error al extraer los parametros {file_name}: {e}")
                results.append((None, None, None, None, None, None, None, None, None, None,None,None,None, None))
        return results
    
class MetallurgicalProcess:
    def __init__(self,df, parameters,Raffinate_Density,PLS_Density):
        self.df = df
        self.parameters = parameters
        self.Raffinate_Density = Raffinate_Density
        self.PLS_Density = PLS_Density
        
    def area_column(self,column_size):
        db_area = {
        'Column_Size': ['4"X8 m', '6"X6 m', '6"X1 m', '6"X3 m', '8"X1 m','8"X4 m','8"X6 m','3mX6 m','8"X8 m','8"X10 m','8"X12 m','8"X16 m','8"X20 m',
                '8"X24 m','8"X30 m','12"X3 m','30"X3 m','1mX10 m','4mX5 m', '6mX1 m'],
        'Column_Diameter(in)': [4.0, 6.0, 6.0, 6.0, 8.0, 8.0 ,8.0, 118.1, 8.0 ,8.0, 8.0, 8.0, 8.0, 8.0, 8.0, 12.0, 30.0, 39.4, 157.5, 236.2]}
        df_area = pd.DataFrame(db_area)
        df_area['Column_Area(m2)'] = math.pi*0.25*(df_area['Column_Diameter(in)']*2.54*0.01)**2
        columna_buscada = 'Column_Size' 
        fila_dato = df_area[df_area[columna_buscada] == column_size].index[0]
        Column_Area = df_area.iloc[fila_dato]['Column_Area(m2)']
        return Column_Area

    def process(self):
        dfs1 = []
        for df, params in zip(self.df, self.parameters):
            try:
                (dry_column_charge, column_size, peso_inicial, cu_inicial, h2so4_inicial, fe_total_inicial, fe_2_inicial, altura_inicial_mineral, cu_total, fe_total, cu_acido,cu_cn, cu_residual,material_moisture) = params
                fe_3_inicial = fe_total_inicial - fe_2_inicial
                #reffinate
                df_feed = df[['Nombre del Archivo','Fecha','Día','Peso(g)_Feed','Cu(g/Kg)_Feed','ÁcidoLibre(g/Kg)_Feed','Fe(g/Kg)_Feed',
                'Fe+2(g/Kg)_Feed','ORP(mV)_Feed','pH_Feed']].fillna(0)
                df_feed['Día'] = pd.to_numeric(df_feed['Día'], errors='coerce')
                df_feed['Fecha'] = pd.to_datetime(df_feed['Fecha'])
                df_feed['Nombre del Archivo'] = df_feed['Nombre del Archivo'].str.replace('.xlsm','')
                df_feed['Feedin_Ratio(g/g)'] = (df_feed['Peso(g)_Feed'].cumsum()) / (dry_column_charge * 1000)
                df_feed['Actual_Rate(Kg/hm2)'] = (df_feed['Peso(g)_Feed'].cumsum()) / (1000 * df_feed['Día'] * 24 * self.area_column(column_size))
                df_feed['Fe+3(g/Kg)_Feed'] = df_feed['Fe(g/Kg)_Feed'] - df_feed['Fe+2(g/Kg)_Feed']
                df_feed['Peso(g)_Feed'].iloc[0] = peso_inicial
                df_feed['Cu(g/Kg)_Feed'].iloc[0] = cu_inicial
                df_feed['ÁcidoLibre(g/Kg)_Feed'][0] = h2so4_inicial
                df_feed['Fe(g/Kg)_Feed'].iloc[0] = fe_total_inicial
                df_feed['Fe+2(g/Kg)_Feed'].iloc[0] = fe_2_inicial
                df_feed['Fe+3(g/Kg)_Feed'].iloc[0] = fe_3_inicial
            except Exception as e:
                print(f"Error en el archivo {df['Nombre del Archivo'].iloc[0]}:{e}")
                continue

            #pls 
            df_pls = df[['Peso(g)_PLS','Cu(g/Kg)_PLS','ÁcidoLibre(g/Kg)_PLS','Fe(g/Kg)_PLS','Fe+2(g/Kg)_PLS','ORP(mV)_PLS','pH_PLS']].fillna(0)
            df_pls['pH_PLS'] = pd.to_numeric(df_pls['pH_PLS'], errors='coerce')
            df_pls['Feedin_Ratio(g/g)_pls'] = (df_pls['Peso(g)_PLS'].cumsum())/(dry_column_charge*1000)
            df_pls['Actual_Rate(Kg/hm2)'] = (df_pls['Peso(g)_PLS'].cumsum())/(1000*df_feed['Día']*24*self.area_column(column_size))
            df_pls['Fe+3(g/Kg)_PLS'] = df_pls['Fe(g/Kg)_PLS'] - df_pls['Fe+2(g/Kg)_PLS']
            df_pls['Cu_Neto(g/Kg)_PLS'] = df_pls['Cu(g/Kg)_PLS'] - df_feed['Cu(g/Kg)_Feed']

            #others
            num_rows = len(df_pls)
            df_others = pd.DataFrame(index=range(num_rows))
            Free_Column_Height = [(int(column_size[:1]) - altura_inicial_mineral) * 100 if i == 0 else 0 for i in range(num_rows)]
            df_others['Raffinate_Density'] = self.Raffinate_Density
            df_others['PLS_Density'] = self.PLS_Density
            df_others['Free_Column_Height'] = Free_Column_Height
            df_others['Cumulated Compaction'] = df_others['Free_Column_Height'].cumsum()
            df_others['Cumulated Compaction'][1:] = df_others['Cumulated Compaction'][1:] - df_others['Cumulated Compaction'][1]
            df_others['Cumulated Compaction'][0] = 0
            df_others

            #raffinate_calculations
            Cu_feed_incial = cu_inicial*peso_inicial/1000
            H2SO4_feed_inicial = dry_column_charge*h2so4_inicial
            Added_H2SO4_Cumulated_inicial = H2SO4_feed_inicial/dry_column_charge
            Fe_Total_feed_inicial = fe_total_inicial*peso_inicial/1000
            Fe_2_feed_inicial = fe_2_inicial*peso_inicial/1000
            Fe_3_feed_inicial = Fe_Total_feed_inicial - Fe_2_feed_inicial
            Added_Fe_3_Cumulated_inicial = Fe_3_feed_inicial/dry_column_charge
            if fe_total_inicial == 0:
                Fe3_Fe2_inicial = 0
            else:
                Fe3_Fe2_inicial = fe_3_inicial/fe_2_inicial

            if fe_total_inicial == 0:
                Fe2_FeT_inicial = 0
            else:
                Fe2_FeT_inicial = fe_2_inicial/fe_total_inicial
            Added_FeT_Cumulated_inicial =  Fe_Total_feed_inicial/dry_column_charge
            df_feed_calc = pd.DataFrame()
            df_feed_calc['Cu(g)'] = df_feed['Peso(g)_Feed']*df_feed['Cu(g/Kg)_Feed']/1000
            df_feed_calc['Cu(g)'][0] = Cu_feed_incial

            df_feed_calc['H2SO4(g)'] = df_feed['Peso(g)_Feed']*df_feed['ÁcidoLibre(g/Kg)_Feed']/1000
            df_feed_calc['H2SO4(g)'][0] = H2SO4_feed_inicial

            df_feed_calc['Added_H2SO4_Cumulated(Kg/Tn)'] = (df_feed_calc['H2SO4(g)'].cumsum() + H2SO4_feed_inicial)/dry_column_charge
            df_feed_calc['Added_H2SO4_Cumulated(Kg/Tn)'][0] = Added_H2SO4_Cumulated_inicial

            df_feed_calc['Fe_Total(g)'] = df_feed['Peso(g)_Feed']*df_feed['Fe(g/Kg)_Feed']/1000
            df_feed_calc['Fe_Total(g)'][0] = Fe_Total_feed_inicial

            df_feed_calc['Fe+2(g)'] = df_feed['Peso(g)_Feed']*df_feed['Fe+2(g/Kg)_Feed']/1000
            df_feed_calc['Fe+2(g)'][0] = Fe_2_feed_inicial

            df_feed_calc['Fe+3(g)'] = df_feed['Peso(g)_Feed']*df_feed['Fe+3(g/Kg)_Feed']/1000
            df_feed_calc['Fe+3(g)'][0] = Fe_3_feed_inicial

            df_feed_calc['AAdded_Fe+3_Cumulated(Kg/Tn)'] = (df_feed_calc['Fe+3(g)'].cumsum() + Fe_3_feed_inicial)/dry_column_charge
            df_feed_calc['AAdded_Fe+3_Cumulated(Kg/Tn)'][0] = Added_Fe_3_Cumulated_inicial

            df_feed_calc['Fe+3/Fe+2'] = (df_feed['Fe+3(g/Kg)_Feed']/df_feed['Fe+2(g/Kg)_Feed']).fillna(0)
            df_feed_calc['Fe+3/Fe+2'][0] = Fe3_Fe2_inicial

            if (df_feed['Peso(g)_Feed'] == 0).all():
                df_feed_calc['Fe+2/FeT'] = 0
            else:
                condicion = (df_feed['Fe+2(g/Kg)_Feed'] <= 0) | (df_feed['Fe+2(g/Kg)_Feed'] > df_feed['Fe(g/Kg)_Feed'])
                df_feed_calc['Fe+2/FeT'] = pd.Series(index=df_feed.index)
                df_feed_calc.loc[condicion, 'Fe+2/FeT'] = 0
                df_feed_calc.loc[~condicion, 'Fe+2/FeT'] = df_feed.loc[~condicion, 'Fe+2(g/Kg)_Feed'] / df_feed.loc[~condicion, 'Fe(g/Kg)_Feed']
            df_feed_calc['Fe+2/FeT'][0] = Fe2_FeT_inicial

            df_feed_calc['Added_FeT_Cumulated(Kg/Tn)'] = (df_feed_calc['Fe_Total(g)'].cumsum() + Fe_Total_feed_inicial)/dry_column_charge
            df_feed_calc['Added_FeT_Cumulated(Kg/Tn)'][0] = Added_FeT_Cumulated_inicial

            #pls calculations
            df_pls_calc = pd.DataFrame()
            df_pls_calc['Cu(g)_pls'] = df_pls['Peso(g)_PLS']*df_pls['Cu(g/Kg)_PLS']/1000
            df_pls_calc['pH*Liters_pls'] = df_pls['Peso(g)_PLS']*df_pls['pH_PLS']
            df_pls_calc['H2SO4(g)_pls'] = df_pls['Peso(g)_PLS']*df_pls['ÁcidoLibre(g/Kg)_PLS']/1000
            df_pls_calc['Fe_Total(g)_pls'] = df_pls['Peso(g)_PLS']*df_pls['Fe(g/Kg)_PLS']/1000
            df_pls_calc['Fe+2(g)_pls'] = df_pls['Peso(g)_PLS']*df_pls['Fe+2(g/Kg)_PLS']/1000
            df_pls_calc['Fe+3(g)_pls'] = df_pls['Peso(g)_PLS']*df_pls['Fe+3(g/Kg)_PLS']/1000
            df_pls_calc['Cu/FeT_pls'] = (df_pls['Cu(g/Kg)_PLS']/df_pls['Fe(g/Kg)_PLS']).fillna(0)
            df_pls_calc['Cu/Fe+2_pls'] = (df_pls['Cu(g/Kg)_PLS']/df_pls['Fe+2(g/Kg)_PLS']).fillna(0)

            condicion_1 = (df_pls['Peso(g)_PLS'] == 0) | (df_pls['Fe+2(g/Kg)_PLS'] > df_pls['Fe(g/Kg)_PLS'])
            df_pls_calc['Fe+3/Fe+2_pls'] = pd.Series(index=df_pls.index)
            df_pls_calc.loc[condicion_1, 'Fe+3/Fe+2_pls'] = 0
            df_pls_calc.loc[~condicion_1, 'Fe+3/Fe+2_pls'] = df_pls.loc[~condicion_1, 'Fe+3(g/Kg)_PLS'] / df_pls.loc[~condicion_1, 'Fe+2(g/Kg)_PLS']

            condicion_2 = (df_pls['Peso(g)_PLS'] == 0) | (df_pls['Fe+2(g/Kg)_PLS'] > df_pls['Fe(g/Kg)_PLS'])
            df_pls_calc['Fe+2/Fe_pls'] = pd.Series(index=df_pls.index)
            df_pls_calc.loc[condicion_2, 'Fe+2/Fe_pls'] = 0
            df_pls_calc.loc[~condicion_2, 'Fe+2/Fe_pls'] = df_pls.loc[~condicion_2, 'Fe+2(g/Kg)_PLS'] / df_pls.loc[~condicion_2, 'Fe(g/Kg)_PLS']

            #element balance
            df_eb = pd.DataFrame()
            if (df_pls_calc['Cu(g)_pls'] == 0).all():
                df_eb['Leached_Cu_day(g)'] = -1*df_feed_calc['Cu(g)']
            else:
                df_eb['Leached_Cu_day(g)'] =  df_pls_calc['Cu(g)_pls'] - df_feed_calc['Cu(g)']
            df_eb['Leached_Cu_day(g)'][0] = 0
            df_eb['Leached_Cu_cumulated(g)'] = df_eb['Leached_Cu_day(g)'].cumsum()
            df_eb['Leached_Cu_cumulated(mol)'] = df_eb['Leached_Cu_cumulated(g)']/63.55

            df_eb['Leached_Fe_day(g)'] = df_pls_calc['Fe_Total(g)_pls'] - df_feed_calc['Fe_Total(g)']
            df_eb['Leached_Fe_day(g)'][0] = 0
            df_eb['Leached_Fe_cumulated(g)'] = df_eb['Leached_Fe_day(g)'].cumsum()
            df_eb['Leached_Fe_cumulated(mol)'] = df_eb['Leached_Fe_cumulated(g)']/dry_column_charge

            if (df_pls_calc['Fe+2(g)_pls'] == 0).all():
                df_eb['Leached_Fe+2_day(g)'] = -1*df_feed_calc['Fe+2(g)']
            else:
                df_eb['Leached_Fe+2_day(g)'] =  df_pls_calc['Fe+2(g)_pls'] - df_feed_calc['Fe+2(g)']
            df_eb['Leached_Fe+2_cumulated(g)'] = df_eb['Leached_Fe+2_day(g)'].cumsum()

            if (df_pls_calc['Fe+3(g)_pls'] == 0).all():
                df_eb['Consumed_Fe+3_day(g)'] = df_feed_calc['Fe+3(g)']
            else:
                df_eb['Consumed_Fe+3_day(g)'] =  - df_pls_calc['Fe+3(g)_pls'] + df_feed_calc['Fe+3(g)']
            df_eb['Consumed_Fe+3_cumulated(g)'] = df_eb['Consumed_Fe+3_day(g)'].cumsum()
            df_eb['Consumed_Fe+3_cumulated(mole)'] = df_eb['Consumed_Fe+3_cumulated(g)']/55.85
            
            df_eb['Consumed_Fe+3_cumulated(Kg/Tn)'] = df_eb['Consumed_Fe+3_cumulated(g)']/dry_column_charge

            condicion_3 = (df_pls['Feedin_Ratio(g/g)_pls'] == 0)
            df_eb['Fe+3_Cu_cumulated(mole/mole)'] = pd.Series(index=df_pls.index)
            df_eb.loc[condicion_3, 'Fe+3_Cu_cumulated(mole/mole)'] = 0
            df_eb.loc[~condicion_3, 'Fe+3_Cu_cumulated(mole/mole)'] = df_eb.loc[~condicion_3, 'Consumed_Fe+3_cumulated(mole)'] / df_eb.loc[~condicion_3, 'Leached_Cu_cumulated(mol)']

            condicion_4 = (df_pls['Feedin_Ratio(g/g)_pls'] == 0)
            df_eb['Fe+3_Cu_cumulated(Kg/Kg)'] = pd.Series(index=df_pls.index)
            df_eb.loc[condicion_4, 'Fe+3_Cu_cumulated(Kg/Kg)'] = 0
            df_eb.loc[~condicion_4, 'Fe+3_Cu_cumulated(Kg/Kg)'] = df_eb.loc[~condicion_4, 'Consumed_Fe+3_cumulated(g)'] / df_eb.loc[~condicion_4, 'Leached_Cu_cumulated(g)']
            
            df_eb['Consumed_H2SO4_GroosDay(g)'] =  df_feed_calc['H2SO4(g)'] - df_pls_calc['H2SO4(g)_pls']
            df_eb['Consumed_H2SO4_GroosDay(g)'][0] = H2SO4_feed_inicial

            df_eb['Consumend_H2SO4_Eq_Acid_Day_Cu(g)'] = 1.54*df_eb['Leached_Cu_day(g)']
            df_eb['Consumend_H2SO4_Eq_Acid_Day_Fe(g)'] = 1.76*df_eb['Leached_Fe_day(g)']
            df_eb['Consumend_H2SO4_NextDay(g)'] = df_eb['Consumed_H2SO4_GroosDay(g)'] - df_eb['Consumend_H2SO4_Eq_Acid_Day_Cu(g)']

            df_eb['Consumed_H2SO4_GrossDay(Kg/Tn)'] = df_eb['Consumed_H2SO4_GroosDay(g)']/dry_column_charge
            df_eb['Consumed_H2SO4_Acid_Day_Cu(Kg/Tn)'] = df_eb['Consumend_H2SO4_Eq_Acid_Day_Cu(g)']/dry_column_charge
            df_eb['Consumed_H2SO4_Acid_Day_Fe(Kg/Tn)'] = df_eb['Consumend_H2SO4_Eq_Acid_Day_Fe(g)']/dry_column_charge
            df_eb['Consumed_H2SO4_NextDay(Kg)'] = df_eb['Consumend_H2SO4_NextDay(g)']/dry_column_charge

            #recovery
            Cu_Total_g = cu_total*dry_column_charge*1000/100
            Fe_Total_g = fe_total*dry_column_charge*1000/100
            df_rec = pd.DataFrame()
            df_rec['Partial(g/Kg)'] = ((df_pls['Peso(g)_PLS']*df_pls['Cu(g/Kg)_PLS']) - (df_feed['Peso(g)_Feed']*df_feed['Cu(g/Kg)_Feed']))/(dry_column_charge*1000)
            df_rec['Partial(g/Kg)'][0] = 0
            df_rec['Cumulated(g/Kg)'] = df_rec['Partial(g/Kg)'].cumsum()
            df_rec['Cumulated(%CuT)'] = (df_rec['Cumulated(g/Kg)']*100)/(10*cu_total)
            df_rec['Cumulated(g/Kg)'][0] = 0
            df_rec['Residue(%CuT)'] = (Cu_Total_g - df_eb['Leached_Cu_day(g)'].cumsum())/(10*dry_column_charge)
            df_rec['Residue(%CuT)'][0] = 0
            df_rec['Partial_FeT%'] = (df_eb['Leached_Fe_day(g)']/Fe_Total_g) *100
            df_rec['Partial_FeT%'][0] = 0
            
            condition = (df_pls['Feedin_Ratio(g/g)_pls'] == 0 )
            df_rec['Cumulated_FeT%'] = pd.Series(index=df_rec.index)
            df_rec.loc[condition,'Cumulated_FeT%'] = 0
            df_rec.loc[~condition,'Cumulated_FeT%'] = df_rec['Partial_FeT%'].cumsum()

            #ADICIONAL
            df_add = pd.DataFrame()
            df_add['KgFe+2/tms_ore'] = df_eb['Leached_Fe+2_cumulated(g)']/dry_column_charge
            df_add['Cu_extraido(Kg)'] =  df_eb['Leached_Cu_cumulated(g)']/1000
            df_add['Cu_adic_Ref(Kg)'] = df_feed_calc['Cu(g)']/1000
            df_add['Cu_acum_adic_Ref(Kg)'] = df_add['Cu_adic_Ref(Kg)'].cumsum()
            df_add['m3'] = (df_feed['Peso(g)_Feed'].iloc[1:].cumsum())/1000
            df_add['m3']= df_add['m3'].fillna(0)
            df_add['Cu_extraido(Kg/m3)'] = (df_add['Cu_extraido(Kg)']/df_add['m3']).fillna(0)

            condicion_5 = (df_add['Cu_extraido(Kg)'] == 0)
            df_add['Kg_Cu_extr/kg_Cu_ad(Kg/Kg)'] = pd.Series(index=df_add.index)
            df_add.loc[condicion_5, 'Kg_Cu_extr/kg_Cu_ad(Kg/Kg)'] = 0
            df_add.loc[~condicion_5, 'Kg_Cu_extr/kg_Cu_ad(Kg/Kg)'] = df_add.loc[~condicion_5, 'Cu_extraido(Kg)'] / df_add.loc[~condicion_5, 'Cu_acum_adic_Ref(Kg)']

            condicion_6 = (df_eb['Leached_Cu_day(g)']==0)
            df_add['KgCu_extr/dia(Kg)'] = pd.Series(index=df_add.index)
            df_add.loc[condicion_6,'KgCu_extr/dia(Kg)'] = df_add['KgCu_extr/dia(Kg)'].shift(1)
            df_add.loc[~condicion_6, 'KgCu_extr/dia(Kg)'] = df_eb.loc[~condicion_6, 'Leached_Cu_day(g)'] / 1000
            df_add['KgCu_extr/dia(Kg)'] = df_add['KgCu_extr/dia(Kg)'].fillna(0)
            
            if (df_feed['Feedin_Ratio(g/g)'] == 0).all():
                df_add['Raffinate_Fe+3/FeT'] = 0
            else:
                condicion_1 = df_feed['Fe+3(g/Kg)_Feed'] <= 0
                condicion_2 = df_feed['Fe+3(g/Kg)_Feed'] > df_feed['Fe(g/Kg)_Feed']
                df_add['Raffinate_Fe+3/FeT'] = np.where(condicion_1, 0, np.where(condicion_2, 0, df_feed['Fe+3(g/Kg)_Feed'] / df_feed['Fe(g/Kg)_Feed']))
            df_add['Raffinate_Fe+3/FeT'][0] = 0


            condicion_1 = df_pls['Feedin_Ratio(g/g)_pls'] == 0
            condicion_2 = df_pls['Fe+3(g/Kg)_PLS'] <= 0
            condicion_3 = df_pls['Fe+3(g/Kg)_PLS'] > df_pls['Fe(g/Kg)_PLS']

            df_add['pls_Fe+3/FeT'] = np.where(condicion_1, 0, np.where(condicion_2, 0, np.where(condicion_3, 1, df_pls['Fe+3(g/Kg)_PLS'] / df_pls['Fe(g/Kg)_PLS'])))

            condicion_7 = (df_feed['Fe+3(g/Kg)_Feed'] == 0)
            df_add['Fe+3pls/Fe+3ref'] = pd.Series(index=df_add.index)
            df_add.loc[condicion_7, 'Fe+3pls/Fe+3ref'] = 0
            df_add.loc[~condicion_7, 'Fe+3pls/Fe+3ref'] = df_pls.loc[~condicion_7, 'Fe+3(g/Kg)_PLS'] / df_feed.loc[~condicion_7, 'Fe+3(g/Kg)_Feed']

            #consumo de acido
            df_ac = pd.DataFrame()
            df_ac['Gross_Acid_Consumption(Kg/Tn)'] = df_eb['Consumed_H2SO4_GrossDay(Kg/Tn)'].cumsum()
            df_ac['Net_Acid_Consumption(Kg/Tn)'] = df_eb['Consumed_H2SO4_NextDay(Kg)'].cumsum()

            condition = (df_rec['Cumulated(%CuT)'] == 0)
            df_ac['Net_Acid_Consumption(Kg/KgCu)'] = pd.Series(index=df_rec.index)
            df_ac.loc[condition, 'Net_Acid_Consumption(Kg/KgCu)'] = 0
            df_ac.loc[~condition, 'Net_Acid_Consumption(Kg/KgCu)'] = df_ac.loc[~condition, 'Net_Acid_Consumption(Kg/Tn)'] / (cu_total * df_rec.loc[~condition, 'Cumulated(%CuT)']) * 10

            #VOLUME BALANCE
            df_vol = pd.DataFrame()
            condition = (df_pls['Feedin_Ratio(g/g)_pls'] == 0)
            df_vol['D/R Ratio'] = pd.Series(index=df_pls.index)
            df_vol.loc[condition, 'D/R Ratio'] = 0
            df_vol.loc[~condition, 'D/R Ratio'] = (df_pls.loc[~condition,'Peso(g)_PLS'].cumsum()) / (df_feed.loc[~condition,'Peso(g)_Feed'].cumsum())
            df_vol['Retained_Solution(%vol/vol)'] = (df_feed['Peso(g)_Feed'].cumsum() - df_pls['Peso(g)_PLS'].cumsum())/(df_feed['Peso(g)_Feed'].cumsum())*100
            df_vol['Retained_Solution(%vol/vol)'][0] = 0
            df_vol['Retained_Solution(g/Kg)']= (df_feed['Peso(g)_Feed'].cumsum() - df_pls['Peso(g)_PLS'].cumsum())/(dry_column_charge*1000)
            df_vol['Retained_Solution(g/Kg)'][0] = 0
            
            df_final_1 = pd.concat([df_feed,df_pls,df_others,df_rec,df_ac,df_eb,df_feed_calc,df_pls_calc,df_vol,df_add],axis=1)

            df_IL = pd.DataFrame()
            db = df_rec['Cumulated(%CuT)']
            df_IL['Ext%Cu_oxidos'] = (db*cu_total)/cu_acido
            df_IL['Ext%Cu_facil'] = (db*cu_total)/(cu_acido + 0.5*cu_cn)
            df_IL['Ext%Cu_IL'] = (db*cu_total)/(cu_acido + cu_cn)
            df_final_2 = pd.concat([df_final_1, df_IL], axis=1)
            df_final_2 = df_final_2.drop(df_final_2.index[0])    
            dfs1.append(df_final_2)
        return dfs1

    def consolidado(self):
        lista_dfs1  = self.process()
        df_consolidado = pd.concat(lista_dfs1,axis=0)
        nuevos_nombres = {'Nombre del Archivo':'columna','Fecha':'fecha','Día':'dias','Peso(g)_Feed':'riego_refino_peso_g','Feedin_Ratio(g/g)':'ratio_riego_refino_m3/t',
                          'Actual_Rate(Kg/hm2)':'ratio_riego_refino_kg/hm2','Cu(g/Kg)_Feed':'Cu_refino_g/kg','ÁcidoLibre(g/Kg)_Feed':'acido_refino_g/kg',
                          'Fe(g/Kg)_Feed':'FeTotal_refino_g/kg','Fe+2(g/Kg)_Feed':'Fe+2_refino_g/kg','Fe+3(g/Kg)_Feed':'Fe+3_refino_g/kg','ORP(mV)_Feed':'ORP_refino_mV',
                          'pH_Feed':'pH_refino','Peso(g)_PLS':'DrainDown_PLS_g','Feedin_Ratio(g/g)_pls':'DrainDownPLS_g/kg','Actual_Rate(Kg/hm2)':'DrainDown_PLS_Kg/hm2',
                          'Cu(g/Kg)_PLS':'Cu_PLS_g/kg','Cu_Neto(g/Kg)_PLS':'CuNet_PLS_g/kg','ÁcidoLibre(g/Kg)_PLS':'acido_PLS_g/kg','Fe(g/Kg)_PLS':'FeTotal_PLS_g/kg',
                          'Fe+2(g/Kg)_PLS':'Fe+2_PLS_g/kg','Fe+3(g/Kg)_PLS':'Fe+3_PLS_g/kg','ORP(mV)_PLS':'ORP_PLS_mV','pH_PLS':'pH_PLS','Raffinate_Density':'densidad_refino_gr/cc',
                          'PLS_Density':'densidad_PLS_gr/cc','Free_Column_Height':'altura_libre_column_cm','Cumulated Compaction':'%_compactacion_column',
                          'Partial(g/Kg)':'ExtParc_CuT_g/kg','Cumulated(g/Kg)':'ExtAcum_CuT_g/kg','Cumulated(%CuT)':'ExtAcum_CuT_%','Residue(%CuT)':'residuo_Cu_acum_%',
                          'Partial_FeT%':'ExtParc_FeT_%','Cumulated_FeT%':'ExtAcum_FeT_%','Gross_Acid_Consumption(Kg/Tn)':'ConsBrutoAci_kg/t','Net_Acid_Consumption(Kg/Tn)':'ConsNeto_Aci_kg/t',
                          'Net_Acid_Consumption(Kg/KgCu)':'ConsNetoAci_kg/kgCu','Cu(g)':'Cu_agregado_refino_g','H2SO4(g)':'acido_parcial_agregado_refino_g',
                          'Added_H2SO4_Cumulated(Kg/Tn)':'acido_acum_agregado_refino_kg/t','Fe_Total(g)':'FeT_agregado_refino_g','Fe+2(g)':'Fe+2_agregado_refino_g',
                          'Fe+3(g)':'Fe+3_agregado_refino_g','AAdded_Fe+3_Cumulated(Kg/Tn)':'Fe+3_acum_agregado_refino_kg/t','Fe+3/Fe+2':'Fe+3/Fe+2_agregado_refino',
                          'Fe+2/FeT':'Fe+2/FeT_agregado_refino','Added_FeT_Cumulated(Kg/Tn)':'FeT_acum_agregado_refino_kg/t','Cu(g)_pls':'Cu_percolado_PLS_g',
                          'pH*Liters_pls':'pH*PLS_percolado','H2SO4(g)_pls':'acido_percolado_PLS_g','Fe_Total(g)_pls':'FeT_percolado_PLS_g','Fe+2(g)_pls':'Fe+2_percolado_PLS_g',
                          'Fe+3(g)_pls':'Fe+3_percolado_PLS_g','Cu/FeT_pls':'CuT/FeT_PLS','Cu/Fe+2_pls':'CuT/Fe+2_PLS','Fe+3/Fe+2_pls':'Fe+3/Fe2+_PLS','Fe+2/Fe_pls':'Fe+2/FeT_PLS',
                          'D/R Ratio':'PLS_percolado/refino_agregado','Retained_Solution(%vol/vol)':'%vol/vol_SolucionRetenida','Retained_Solution(g/Kg)':'SolucionRetenida_g/kg','Dynamic_Moisture':'humedad_dinamica_%',
                          'Leached_Cu_day(g)':'Cu_lixiviado_dia_g','Leached_Cu_cumulated(g)':'Cu_lixiviado_acum_g','Leached_Cu_cumulated(mol)':'Cu_lixiviado_acum_mol','Leached_Fe_day(g)':'Fe_lixiviado_dia_g',
                          'Leached_Fe_cumulated(g)':'Fe_lixiviado_acum_g','Leached_Fe_cumulated(mol)':'Fe_lixiviado_acum_kg/t','Leached_Fe+2_day(g)':'Fe+2_lixiviado_dia_g',
                          'Leached_Fe+2_cumulated(g)':'Fe+2_lixiviado_acum_g','Consumed_Fe+3_day(g)':'Fe+3_consumido_dia_g','Consumed_Fe+3_cumulated(g)':'Fe+3_consumido_acum_g','Consumed_Fe+3_cumulated(mole)':'Fe+3_consumido_acum_mol',
                          'Consumed_Fe+3_cumulated(Kg/Tn)':'Fe+3_consumido_acum/t','Fe+3_Cu_cumulated(mole/mole)':'Fe+3_consumido_acum_mol/Cu_lixiviado_acum_mol','Fe+3_Cu_cumulated(Kg/Kg)':'Fe+3_consumido_acum_kg/Cu_lixiviado_acum_kg',
                          'Consumed_H2SO4_GroosDay(g)':'consumo_bruto_acido_dia_g','Consumend_H2SO4_Eq_Acid_Day_Cu(g)':'consumo_equivalente_dia_acido_Cu_g','Consumend_H2SO4_Eq_Acid_Day_Fe(g)':'consumo_equivalente_dia_acido_Fe_g',
                          'Consumend_H2SO4_NextDay(g)':'consumo_neto_acido_g','Consumed_H2SO4_GrossDay(Kg/Tn)':'consumo_bruto_acido_dia_kg/t','Consumed_H2SO4_Acid_Day_Cu(Kg/Tn)':'consumo_equivalente_dia_acido_Cu_kg/t',
                          'Consumed_H2SO4_Acid_Day_Fe(Kg/Tn)':'consumo_equivalente_dia_acido_Fe_kg/t','Consumed_H2SO4_NextDay(Kg)':'consumo_neto_dia_acido_kg/t','KgFe+2/tms_ore':'Fe+2_lixiviado_acum_kg/t',
                          'Cu_extraido(Kg)':'Cu_extraido_acum_kg','Cu_adic_Ref(Kg)':'Cu_adicionado_refino_dia_kg','Cu_acum_adic_Ref(Kg)':'Cu_adicionado_refino_acum_kg','m3':'m3_agregados_refino_acum',
                          'Cu_extraido(Kg/m3)':'Cu_extraido_kg/refino_agregado_m3','Kg_Cu_extr/kg_Cu_ad(Kg/Kg)':'Cu_extraido_kg/Cu_adicionado_refino','KgCu_extr/dia(Kg)':'Cu_extraido_dia_kg','Raffinate_Fe+3/FeT':'Fe+3/FeT_refino',
                          'pls_Fe+3/FeT':'Fe+3/FeT_PLS','Fe+3pls/Fe+3ref':'Fe+3_PLS / Fe+3_refino','Ext%Cu_oxidos':'Ext%Cu_oxidos','Ext%Cu_facil':'Ext%Cu_facil','Ext%Cu_IL':'Ext%Cu_IL'}
        df_consolidado = df_consolidado.rename(columns=nuevos_nombres)
        return df_consolidado




