import openpyxl
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.legend import Legend

class curvas_cineticas:
    # Corrige los parámetros, se agrega self como primer parámetro
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

    # Agrega self como primer parámetro en las funciones de la clase
    def curvas(self, sheet_graphs_name):
        # Cargar el archivo de Excel
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet_graphs = workbook.create_sheet(sheet_graphs_name)
        
        # Configurar las posiciones iniciales
        start_row = 11
        start_col = 2  # Columna B

        # Obtener el número de filas
        num_rows = sheet.max_row

        # Configurar la posición inicial para los gráficos en la hoja "graficas_columnas"
        graphs_per_row = 3
        graph_height = 15  # Altura de la celda para cada gráfico
        graph_width = 8    # Ancho de la celda para cada gráfico

        # Crear gráficos de dispersión para cada conjunto de datos
        graph_index = 0
        for col_offset in range(0, sheet.max_column, 10):
            if sheet.cell(row=start_row, column=start_col + col_offset).value is None:
                break

            chart = ScatterChart()

            # Obtener el título del gráfico según el valor de una celda específica
            title_row = start_row + 1
            title_col = start_col + col_offset - 1
            chart.title = f'{sheet.cell(row=title_row, column=title_col).value}'

            chart.style = 13
            chart.x_axis.title = 'Días'
            chart.y_axis.title = 'Recuperación Cu, %'

            # Configurar la leyenda para que esté fuera de la gráfica
            legend = Legend()
            legend.position = "r"
            legend.overlay = False
            chart.legend = legend

            # Referencias para los días, % CuT Acumul y Mod Rec Cu
            xvalues = Reference(sheet, min_col=start_col + col_offset, min_row=start_row + 1, max_row=num_rows)
            yvalues1 = Reference(sheet, min_col=start_col + col_offset + 2, min_row=start_row + 1, max_row=num_rows)
            yvalues2 = Reference(sheet, min_col=start_col + col_offset + 3, min_row=start_row + 1, max_row=num_rows)
            yvalues3 = Reference(sheet, min_col=start_col + col_offset + 6, min_row=start_row + 1, max_row=num_rows)

            # Serie 1: Días vs % CuT Acumul (Dispersión)
            series1 = Series(values=yvalues1, xvalues=xvalues, title="% CuT Acumul")
            series1.marker.symbol = "circle"
            series1.marker.graphicalProperties.solidFill = "FF0000"  # Rojo
            series1.marker.graphicalProperties.line.solidFill = "FF0000"  # Blanco
            series1.graphicalProperties.line.noFill = True
            chart.series.append(series1)

            # Crear un gráfico de línea para las series 2 y 3
            line_chart = LineChart()

            # Serie 2: Días vs Mod Rec Cu (Línea)
            series2 = Series(values=yvalues2, xvalues=xvalues, title="Mod Rec Cu")
            series2.graphicalProperties.line.solidFill = "0000FF"  # Azul
            line_chart.series.append(series2)

            # Serie 3: Días vs Mod Rec Cu 2 (Línea)
            series3 = Series(values=yvalues3, xvalues=xvalues, title="Mod Rec Cu 2")
            series3.graphicalProperties.line.solidFill = "00FF00"  # Verde
            line_chart.series.append(series3)

            # Ajustar el rango del eje Y
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100

            # Añadir las series de línea al gráfico de dispersión
            for series in line_chart.series:
                chart.series.append(series)

            # Mostrar los valores en los ejes X e Y
            chart.x_axis.majorTickMark = 'in'
            chart.x_axis.minorTickMark = 'in'
            chart.y_axis.majorTickMark = 'in'
            chart.y_axis.minorTickMark = 'in'

            # Añadir líneas de cuadrícula y personalizar ejes
            chart.x_axis.majorGridlines = ChartLines()
            chart.y_axis.majorGridlines = ChartLines()

            # Configuración del layout del gráfico
            chart.layout = Layout(
                manualLayout=ManualLayout(
                    x=0.05,  # Posición x de la gráfica
                    y=0.05,  # Posición y de la gráfica
                    h=0.8,   # Altura de la gráfica
                    w=0.8    # Anchura de la gráfica
                )
            )

            # Determinar la posición del gráfico en la hoja "graficas_columnas"
            row_pos = (graph_index // graphs_per_row) * graph_height + 1
            col_pos = (graph_index % graphs_per_row) * graph_width + 1
            cell_position = f'{openpyxl.utils.get_column_letter(col_pos)}{row_pos}'

            sheet_graphs.add_chart(chart, cell_position)

            # Incrementar el índice del gráfico
            graph_index += 1

        # Guardar el archivo Excel con los gráficos añadidos
        workbook.save(self.file)
